"""
teamDashboard_v2.py
Reads Excel source files, writes data.json + config.json (if missing).
No Google Sheets dependency. Designed for GitHub Pages + static JSON hosting.

Usage:
  python teamDashboard_v2.py

Output files (written to OUTPUT_DIR):
  data.json    — live dashboard data, updated every run
  config.json  — manually edited monthly (plan $, team name, etc.)
                 created on first run with defaults; never overwritten after that
"""

import os
import sys
import re
import json
from datetime import datetime, date
from calendar import month_abbr
from collections import defaultdict
import openpyxl

# ===================== CONFIG =====================
# Edit these paths to match your machine.

SOURCE_FILES = {
    "lm_quotes": r"C:\Users\jwagemd\OneDrive - Johnson Controls\Documents\Team Dashboard Files\Excel Reports\L&M Quoted Count & Margin Summary Report.xlsx",
    "prejob":    r"C:\Users\jwagemd\OneDrive - Johnson Controls\Documents\Team Dashboard Files\Excel Reports\PreJob Checklist.xlsx",
    "nearmiss":  r"C:\Users\jwagemd\OneDrive - Johnson Controls\Documents\Team Dashboard Files\Excel Reports\Near Miss Summary Report.xlsx",
    "crc":       r"C:\Users\jwagemd\OneDrive - Johnson Controls\Documents\Team Dashboard Files\Excel Reports\CRC Summary Report.xlsx",
    "debrief":   r"C:\Users\jwagemd\OneDrive - Johnson Controls\Documents\Team Dashboard Files\Excel Reports\Debriefing.xlsx",
    "roster":    r"C:\Users\jwagemd\OneDrive - Johnson Controls\Documents\Team Dashboard Files\Excel Reports\Roster.xlsx",
}

# Where to write data.json and config.json.
# Set this to your local git repo folder.
OUTPUT_DIR = r"C:\Users\jwagemd\OneDrive - Johnson Controls\Documents\Team Dashboard Files\repo"

# Team / branch filter constants (used to filter multi-branch exports)
TEAM1          = '0N48 TB Team 1 Denver SVC'
TEAM2          = '0N48 TB Team 2 Denver SVC'
BRANCH_DENVER  = 'DENVER CO CB - 0N48'

# ===================== LOGGING =====================

def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

# ===================== FILE LOADING =====================

def load_wb(key: str):
    """Load an openpyxl workbook by source key. Returns None on failure."""
    path = SOURCE_FILES.get(key)
    if not path:
        log(f"WARNING: No path configured for '{key}'")
        return None
    if not os.path.exists(path):
        log(f"WARNING: File not found — {path}")
        return None
    # Siebel sometimes exports HTML-wrapped XML with an .xlsx extension.
    # Try openpyxl first; fall back to pandas/xlrd if the zip structure is broken.
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception as e1:
        log(f"WARNING: openpyxl failed on '{key}' ({e1}), trying xlrd fallback...")
        try:
            import pandas as pd
            df = pd.read_excel(path, engine='xlrd')
            log(f"  xlrd fallback succeeded for '{key}' — returning DataFrame wrapper")
            return _DataFrameWorkbook(df)
        except Exception as e2:
            log(f"ERROR: Could not open '{key}' with any engine: {e2}")
            return None


class _DataFrameSheet:
    """Minimal openpyxl-like worksheet backed by a pandas DataFrame.
    Lets the parse functions work unchanged when xlrd fallback fires."""
    def __init__(self, df):
        import pandas as pd
        self._df = df.fillna('')

    def iter_rows(self, min_row=1, max_row=None, values_only=False):
        rows = []
        if min_row == 1:
            rows.append(tuple(str(c) for c in self._df.columns))
        for _, row in self._df.iterrows():
            rows.append(tuple(row))
        for r in rows:
            yield r


class _DataFrameWorkbook:
    def __init__(self, df):
        self.active = _DataFrameSheet(df)

# ===================== DATE HELPERS =====================

def current_month_str(dt=None) -> str:
    dt = dt or datetime.now()
    return f"{dt.year} / {dt.month:02d}"

def prev_month_str(dt=None) -> str:
    dt = dt or datetime.now()
    y, m = dt.year, dt.month
    if m == 1:
        return f"{y-1} / 12"
    return f"{y} / {m-1:02d}"

def parse_month_cell(v) -> str:
    """Normalise any month-ish cell value to 'YYYY / MM' string."""
    if v is None:
        return ''
    s = str(v).strip()
    if len(s) >= 7 and s[4:7] == ' / ':
        try:
            y = int(s[:4]); m = int(s[7:9])
            return f"{y} / {m:02d}"
        except Exception:
            pass
    if isinstance(v, (datetime, date)):
        return f"{v.year} / {v.month:02d}"
    try:
        parts = s.replace(',', ' ').split()
        if len(parts) >= 2:
            mon_txt = parts[0][:3].title()
            y = int([p for p in parts if p.isdigit()][0])
            m = list(month_abbr).index(mon_txt)
            return f"{y} / {m:02d}"
    except Exception:
        pass
    try:
        hit = re.search(r'(\d{1,2})[^\d]+(\d{4})', s)
        if hit:
            return f"{int(hit.group(2))} / {int(hit.group(1)):02d}"
    except Exception:
        pass
    return ''

# ===================== NAME NORMALISATION =====================

INVISIBLE = re.compile(r'[\u00A0\u200B\u200C\u200D]')
NICKNAME_MAP = {
    'alex':'alexander','andy':'andrew','ben':'benjamin','bill':'william',
    'bob':'robert','brad':'bradley','chris':'christopher','dan':'daniel',
    'dave':'david','don':'donald','jack':'john','jim':'james','joe':'joseph',
    'johnny':'john','josh':'joshua','kate':'katherine','kathy':'katherine',
    'ken':'kenneth','kev':'kevin','liz':'elizabeth','matt':'matthew',
    'mike':'michael','nick':'nicholas','pat':'patrick','rick':'richard',
    'rob':'robert','sam':'samuel','steve':'steven','tom':'thomas',
}

def normalize_name(s) -> str:
    return str(s or '').strip()

def clean_name_str(s) -> str:
    s = str(s or '')
    s = INVISIBLE.sub(' ', s)
    s = s.replace('\t', ' ')
    s = re.sub(r'[.,/\\-]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def header_map(ws) -> dict:
    headers = [str(c or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    return {h: i for i, h in enumerate(headers)}

def _split_roster_name(full: str):
    parts = (full or '').split(',', 1)
    last = clean_name_str(parts[0]).lower().strip() if parts else ''
    first_full = clean_name_str(parts[1]).lower().strip() if len(parts) == 2 else ''
    first_tok = first_full.split()[0] if first_full else ''
    return last, first_tok

def match_to_roster(emp_raw: str, roster: dict):
    s = clean_name_str(emp_raw).lower()
    if not s:
        return None
    tokens = s.split()
    def canon(tok):
        return NICKNAME_MAP.get(tok, tok)
    cand_first = {canon(tokens[0])} if tokens else set()
    if len(tokens) > 1:
        cand_first.add(canon(tokens[1]))
    comma_style = ',' in (emp_raw or '')
    candidates = []
    for full in roster:
        last_r, first_tok = _split_roster_name(full)
        if not last_r or not first_tok:
            continue
        first_norm = canon(first_tok)
        last_in  = (last_r in tokens) or (tokens and tokens[-1] == last_r)
        first_in = (first_tok in tokens) or (first_norm in cand_first)
        score = (2 if last_in else 0) + (2 if first_in else 0) + (1 if comma_style and last_in and first_in else 0)
        if score > 0:
            candidates.append((score, full))
    if candidates:
        return max(candidates, key=lambda x: x[0])[1]
    # unique last-name fallback
    last_token = tokens[-1] if tokens else ''
    if not last_token:
        return None
    matches = [f for f in roster if _split_roster_name(f)[0] == last_token]
    return matches[0] if len(matches) == 1 else None

# ===================== ROSTER =====================

def get_roster() -> dict:
    """
    Read roster from Roster.xlsx.
    Expected columns (row 3 onward): Full Name | Display Name | Active (yes/no)
    Falls back to an empty dict on failure.
    """
    wb = load_wb('roster')
    if not wb:
        log("WARNING: No roster file — returning empty roster")
        return {}
    ws = wb.active
    roster = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        full = str(row[0]).strip()
        display = str(row[1]).strip() if len(row) > 1 else ''
        active_raw = str(row[2]).strip().lower() if len(row) > 2 else 'yes'
        active = active_raw in ('yes', 'y', 'true', '1')
        if full and display and active and full != 'Add/remove techs here.':
            roster[full] = display
    log(f"Roster loaded: {len(roster)} techs")
    return roster

# ===================== QUOTES =====================

def parse_quotes(roster: dict, month_str: str) -> list:
    wb = load_wb('lm_quotes')
    if not wb:
        return []
    ws = wb.active
    h = header_map(ws)
    need = ['Quote Month', 'Employee Name', 'Price After Discount', 'Quote Name']
    missing = [k for k in need if k not in h]
    if missing:
        log(f"Quotes: missing columns {missing}")
        return []
    idx = {k: h[k] for k in need}
    agg = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not any(r):
            continue
        if parse_month_cell(r[idx['Quote Month']]) != month_str:
            continue
        matched = match_to_roster(normalize_name(r[idx['Employee Name']]), roster)
        if not matched:
            continue
        qid = str(r[idx['Quote Name']] or '').strip()
        rev = float(r[idx['Price After Discount']] or 0)
        node = agg.setdefault(matched, {'quotes': set(), 'revenue': 0.0})
        if qid:
            node['quotes'].add(qid)
        node['revenue'] += rev
    out = []
    for full, v in agg.items():
        out.append({
            'full_name':    full,
            'display_name': roster[full],
            'month':        month_str,
            'quotes':       len(v['quotes']),
            'revenue':      round(v['revenue'], 2),
        })
    log(f"Quotes parsed: {len(out)} records for {month_str}")
    return out

# ===================== AWARDS (previous month) =====================

def compute_awards(roster: dict) -> dict:
    target = prev_month_str()
    wb = load_wb('lm_quotes')
    empty = {'prev_month': target, 'by_quotes': ('', 0), 'by_dollars': ('', 0.0)}
    if not wb:
        return empty
    ws = wb.active
    h = header_map(ws)
    need = ['Quote Month', 'Employee Name', 'Price After Discount']
    if any(k not in h for k in need):
        return empty
    idx = {k: h[k] for k in need}
    agg_q = defaultdict(int)
    agg_d = defaultdict(float)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not any(r):
            continue
        if parse_month_cell(r[idx['Quote Month']]) != target:
            continue
        matched = match_to_roster(normalize_name(r[idx['Employee Name']]), roster)
        if not matched:
            continue
        dn = roster[matched]
        agg_q[dn] += 1
        agg_d[dn] += float(r[idx['Price After Discount']] or 0)
    by_q = max(agg_q.items(), key=lambda x: (x[1], x[0])) if agg_q else ('', 0)
    by_d = max(agg_d.items(), key=lambda x: (x[1], x[0])) if agg_d else ('', 0.0)
    log(f"Awards computed for {target}: quotes→{by_q[0]}, dollars→{by_d[0]}")
    return {'prev_month': target, 'by_quotes': by_q, 'by_dollars': by_d}

# ===================== SAFETY =====================

def parse_safety(roster: dict, month_str: str, unmatched: list) -> list:
    wb_nm  = load_wb('nearmiss')
    wb_pre = load_wb('prejob')

    # Near miss lookup: full_name → bool
    nm_done = {}
    if wb_nm:
        ws_nm = wb_nm.active
        h_nm = header_map(ws_nm)
        i_emp  = h_nm.get('Employee')
        i_done = h_nm.get('NM Completed')
        for r in ws_nm.iter_rows(min_row=2, values_only=True):
            if not r or i_emp is None:
                continue
            matched = match_to_roster(normalize_name(r[i_emp]).lstrip(), roster)
            if not matched:
                unmatched.append(str(r[i_emp] or '').strip())
                continue
            nm_done[matched] = bool(i_done is not None and int(r[i_done] or 0) >= 1)

    if not wb_pre:
        return []
    ws_pre = wb_pre.active
    h = header_map(ws_pre)
    need = ['Employee', 'PJC Required', 'PJC Completed', '% PJC']
    missing = [k for k in need if k not in h]
    if missing:
        log(f"Safety: missing columns {missing}")
        return []
    i_emp, i_req, i_comp, i_pct = h['Employee'], h['PJC Required'], h['PJC Completed'], h['% PJC']
    data = []
    for r in ws_pre.iter_rows(min_row=2, values_only=True):
        if not r or not any(r):
            continue
        matched = match_to_roster(normalize_name(r[i_emp]), roster)
        if not matched:
            unmatched.append(str(r[i_emp] or '').strip())
            continue
        pct_raw = r[i_pct] if i_pct is not None else 0
        try:
            pct = float(pct_raw or 0.0)
        except Exception:
            pct = 0.0
        pct = pct if pct <= 1 else pct / 100.0
        data.append({
            'full_name':    matched,
            'display_name': roster[matched],
            'month':        month_str,
            'required':     int(r[i_req] or 0),
            'completed':    int(r[i_comp] or 0),
            'pct':          round(pct, 4),
            'near_miss':    nm_done.get(matched, False),
        })
    log(f"Safety parsed: {len(data)} records")
    return data

# ===================== DEBRIEFING =====================

def _to_pct(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        x = float(v)
        return x if x <= 1 else x / 100.0
    s = re.sub(r'[^0-9.\-]', '', str(v))
    if not s:
        return 0.0
    try:
        x = float(s)
        return x if x <= 1 else x / 100.0
    except Exception:
        return 0.0

def parse_debriefing(roster: dict, unmatched: list) -> dict:
    wb = load_wb('debrief')
    if not wb:
        return {}
    ws = wb.active
    h = header_map(ws)
    name_col = h.get('Technician') or h.get('Employee') or h.get('Tech')
    pct_col  = h.get('Tech On Time Debrief %')
    i_branch = h.get('Branch')
    i_ops    = h.get('Operations Team Name')
    if name_col is None or pct_col is None:
        log("Debriefing: missing required columns — skipping")
        return {}
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not any(r):
            continue
        if i_branch is not None:
            br = str(r[i_branch] or '').strip()
            if br and br != BRANCH_DENVER:
                continue
        if i_ops is not None:
            ot = str(r[i_ops] or '').strip()
            if ot and ot not in (TEAM1, TEAM2):
                continue
        emp = normalize_name(r[name_col])
        matched = match_to_roster(emp, roster)
        if not matched:
            unmatched.append(emp)
            continue
        out[matched] = round(_to_pct(r[pct_col]), 4)
    log(f"Debrief parsed: {len(out)} matched")
    return out

# ===================== REVENUE =====================

def parse_crc_revenue() -> dict:
    wb = load_wb('crc')
    if not wb:
        return {'team1': 0.0, 'team2': 0.0, 'total': 0.0}
    ws = wb.active
    h = header_map(ws)
    if 'Operations Team' not in h or 'Revenue' not in h:
        log("CRC: missing 'Operations Team' or 'Revenue' column")
        return {'team1': 0.0, 'team2': 0.0, 'total': 0.0}
    i_team, i_rev = h['Operations Team'], h['Revenue']
    t1 = t2 = 0.0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not any(r):
            continue
        team = str(r[i_team] or '').strip()
        rev  = float(r[i_rev] or 0.0)
        if team == TEAM1:
            t1 = rev
        elif team == TEAM2:
            t2 = rev
    log(f"CRC Revenue — T1: ${t1:,.2f}  T2: ${t2:,.2f}")
    return {'team1': t1, 'team2': t2, 'total': round(t1 + t2, 2)}

# ===================== CONFIG.JSON =====================

def load_or_create_config(config_path: str) -> dict:
    """
    Load config.json if it exists. If not, create it with defaults and
    print instructions. Never overwrites an existing config.
    """
    defaults = {
        "_instructions": {
            "plan": "Update 'denver_plan' at the start of each month with the monthly revenue target ($).",
            "team_name": "Display name shown in the dashboard header.",
            "branch": "Used for display only."
        },
        "team_name": "Team 2",
        "branch": "Denver CO - 0N48",
        "denver_plan": 0
    }
    if not os.path.exists(config_path):
        with open(config_path, 'w') as f:
            json.dump(defaults, f, indent=2)
        log(f"Created config.json at {config_path} — edit 'denver_plan' before next run.")
        return defaults
    with open(config_path) as f:
        cfg = json.load(f)
    log(f"Config loaded — plan: ${cfg.get('denver_plan', 0):,}")
    return cfg

# ===================== ASSEMBLE + WRITE data.json =====================

def build_output(month_str, roster, quotes, safety, debrief, crc_vals, awards, config) -> dict:
    q_map = {r['display_name']: r for r in quotes}
    s_map = {r['display_name']: r for r in safety}

    techs = []
    for full, display in roster.items():
        q = q_map.get(display, {})
        s = s_map.get(display, {})
        pct = s.get('pct', 0.0)
        pct = pct if pct <= 1 else pct / 100.0
        techs.append({
            "name":       display,
            "quotes":     q.get('quotes', 0),
            "revenue":    q.get('revenue', 0.0),
            "safetyPct":  round(pct, 4),
            "required":   s.get('required', 0),
            "completed":  s.get('completed', 0),
            "nearMiss":   s.get('near_miss', False),
            "debriefPct": round(debrief.get(full, 0.0), 4),
        })

    return {
        "generated":  datetime.now().isoformat(timespec='seconds'),
        "month":      month_str,
        "teamName":   config.get('team_name', 'Team 2'),
        "branch":     config.get('branch', ''),
        "techs":      techs,
        "revenue": {
            "team1":  crc_vals['team1'],
            "team2":  crc_vals['team2'],
            "total":  crc_vals['total'],
            "plan":   float(config.get('denver_plan', 0)),
        },
        "awards": {
            "prevMonth":  awards['prev_month'],
            "byQuotes":  {"name": awards['by_quotes'][0],  "count":   awards['by_quotes'][1]},
            "byDollars": {"name": awards['by_dollars'][0], "dollars": round(float(awards['by_dollars'][1]), 2)},
        },
    }

# ===================== MAIN =====================

def main():
    month_str = current_month_str()
    log(f"Starting sync — {month_str}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    config_path = os.path.join(OUTPUT_DIR, 'config.json')
    data_path   = os.path.join(OUTPUT_DIR, 'data.json')

    config  = load_or_create_config(config_path)
    roster  = get_roster()
    unmatched = []

    quotes  = parse_quotes(roster, month_str)
    awards  = compute_awards(roster)
    safety  = parse_safety(roster, month_str, unmatched)
    crc     = parse_crc_revenue()
    debrief = parse_debriefing(roster, unmatched)

    if unmatched:
        unique_unmatched = sorted(set(x for x in unmatched if str(x).strip()))
        log(f"Unmatched names ({len(unique_unmatched)}): {unique_unmatched}")

    output = build_output(month_str, roster, quotes, safety, debrief, crc, awards, config)

    with open(data_path, 'w') as f:
        json.dump(output, f, indent=2)

    log(f"data.json written → {data_path}")
    log("Sync complete.")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        log(f"FATAL: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
